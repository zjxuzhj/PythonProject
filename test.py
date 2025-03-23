from openpyxl import Workbook

# 创建工作簿和工作表
wb = Workbook()
ws = wb.active

# 写入表头和数据
ws['A1'] = "名称"
ws['B1'] = "编号"
ws['C1'] = "成交量"
ws['D1'] = "涨幅"
data = [("比亚迪", "002594","106亿","-7%"), ("东方财富", "300059","77.78亿","-2.27%")]
for row in data:
    ws.append(row)

# 保存文件
wb.save("sales_report.xlsx")